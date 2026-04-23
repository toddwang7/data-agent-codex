from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any, Iterable

from openpyxl import load_workbook


DATE_RANGE_RE = re.compile(r"(?P<start>\d{4}-\d{2}-\d{2})至(?P<end>\d{4}-\d{2}-\d{2})")
FILENAME_MONTH_RE = re.compile(r"(?P<year>20\d{2})(?P<month>\d{2})")


@dataclass(frozen=True)
class SheetSnapshot:
    headers: list[str]
    rows: list[tuple[Any, ...]]
    max_row: int
    max_col: int


def inspect_workbook(path: Path, sample_size: int = 50) -> SheetSnapshot:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    iterator = worksheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(iterator)]
    rows = []
    for row in iterator:
        rows.append(tuple(row))
        if len(rows) >= sample_size:
            break
    return SheetSnapshot(
        headers=headers,
        rows=rows,
        max_row=worksheet.max_row,
        max_col=worksheet.max_column,
    )


def map_headers(headers: Iterable[str], mapping_hints: list[dict[str, Any]]) -> dict[str, Any]:
    headers_list = list(headers)
    hint_by_header = {hint["source_header"]: hint for hint in mapping_hints}
    recognized = []
    unmapped = []
    for header in headers_list:
        hint = hint_by_header.get(header)
        if hint:
            recognized.append(
                {
                    "source_header": header,
                    "target_field": hint["target_field"],
                    "confidence": hint["confidence"],
                }
            )
        else:
            unmapped.append(header)
    return {
        "recognized": recognized,
        "unmapped": unmapped,
        "recognized_count": len(recognized),
        "total_count": len(headers_list),
    }


def _parse_date(value: str) -> date | None:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_date_range(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    text = str(value).strip()
    match = DATE_RANGE_RE.fullmatch(text)
    if not match:
        return None
    start = _parse_date(match.group("start"))
    end = _parse_date(match.group("end"))
    if not start or not end:
        return None
    return {
        "raw": text,
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "same_month": start.year == end.year and start.month == end.month,
        "month_key": f"{start.year:04d}-{start.month:02d}" if start.year == end.year and start.month == end.month else None,
    }


def infer_reporting_months(file_path: Path, parsed_ranges: list[dict[str, Any]]) -> dict[str, Any]:
    candidates: list[dict[str, Any]] = []

    match = FILENAME_MONTH_RE.search(file_path.stem)
    if match:
        filename_month = f"{match.group('year')}-{match.group('month')}"
        candidates.append(
            {
                "source": "file_name_month_tag",
                "value": filename_month,
                "confidence": 0.95,
            }
        )

    month_counter: Counter[str] = Counter()
    for item in parsed_ranges:
        month_key = item.get("month_key")
        if month_key:
            month_counter[month_key] += 1
    for month_key, count in month_counter.most_common(3):
        candidates.append(
            {
                "source": "dominant_time_range_month",
                "value": month_key,
                "confidence": round(count / max(1, len(parsed_ranges)), 3),
                "matched_rows": count,
            }
        )

    default = candidates[0]["value"] if candidates else None
    return {"default": default, "candidates": candidates}


def classify_special_samples(
    rows: list[dict[str, Any]],
    text_rule: dict[str, Any],
    limit: int = 10,
) -> list[dict[str, Any]]:
    fields = text_rule.get("fields", [])
    priority_keywords = text_rule.get("priority_keywords", {})
    results = []
    for row in rows:
        texts = [str(row.get(field) or "") for field in fields]
        merged = " ".join(texts)
        if not merged.strip():
            continue
        classification = "normal"
        matched_keywords: list[str] = []
        for label in ("exclude", "include_with_warning"):
            keywords = priority_keywords.get(label, [])
            hit = [keyword for keyword in keywords if keyword in merged]
            if hit:
                classification = label
                matched_keywords = hit
                break
        if classification == "normal":
            continue
        results.append(
            {
                "classification": classification,
                "matched_keywords": matched_keywords,
                "adcode": row.get("adcode"),
                "remark": row.get("remark"),
                "adslot": row.get("adslot"),
            }
        )
        if len(results) >= limit:
            break
    return results


def extract_row_dicts(
    file_path: Path,
    mapping_hints: list[dict[str, Any]],
    limit: int | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    iterator = worksheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(iterator)]
    header_to_index = {header: index for index, header in enumerate(headers)}
    source_field_map = {
        hint["target_field"]: hint["source_header"]
        for hint in mapping_hints
    }
    row_dicts = []
    for row in iterator:
        row_dict: dict[str, Any] = {}
        for target_field, source_header in source_field_map.items():
            index = header_to_index.get(source_header)
            if index is not None:
                row_dict[target_field] = row[index]
        row_dicts.append(row_dict)
        if limit is not None and len(row_dicts) >= limit:
            break
    return headers, row_dicts
